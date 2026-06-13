from pathlib import Path
import openpyxl, csv

# Get all excel files in the folder.
path = Path(__file__)
excels = [file for file in path.parent.iterdir() if file.is_file() and file.suffix == '.xlsx']
print(f"Total {len(excels)} xlsx files.")

# Create a folder to store csv files
csv_folder = path.parent/'csv'
csv_folder.mkdir(exist_ok=True)

# Loop the excel files and open the xlsx file
for excel_file in excels:
    wb = openpyxl.load_workbook(excel_file, data_only=True)
    file_name = excel_file.stem
    
    # Loop sheets in the workbook
    for sheet_name in wb.sheetnames:
        
        # Create the new name to write into csv
        csv_name = csv_folder/f"{file_name}_{sheet_name}.csv"
        print(f"Writing {excel_file.name}[{sheet_name}]to {csv_name.name}...")

        # Open a csv file ready to write row
        with open(csv_name, 'w', newline='', encoding='utf-8') as csv_file:
            writer = csv.writer(csv_file)
            # Get each sheet
            sheet = wb[sheet_name]
            # Get sheet each row's values as a list
            for row_values in sheet.iter_rows(values_only=True):
                # write to csv files
                writer.writerow(row_values)
        
            
       
print("Done.")




        